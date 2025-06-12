from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
import uuid

EXCEL_PATH = "visitor_data.xlsx"

def save_to_excel(gender, age, ad_image_path, face_path, timestamp):
    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.append(["Timestamp", "Gender", "Age", "Ad Image", "Face Image"])
        wb.save(EXCEL_PATH)

    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    # Save image to unique path
    unique_face_filename = f"{str(uuid.uuid4())}.jpg"
    saved_face_path = os.path.join("faces", unique_face_filename)
    os.rename(face_path, saved_face_path)  # Rename saved image uniquely

    row = [timestamp, gender, age, ad_image_path, saved_face_path]
    ws.append(row)

    # Attach face image to Excel file (optional but useful)
    img = ExcelImage(saved_face_path)
    img.width = 80
    img.height = 80
    ws.add_image(img, f"F{ws.max_row}")

    wb.save(EXCEL_PATH)
